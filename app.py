from flask import Flask, render_template, request, jsonify, send_file
import json
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import psycopg2
from psycopg2.extras import RealDictCursor

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

HEADERS = [
    "ქალაქი (ლოკაცია)", "სახელი გვარი", "პირადი #", "ბრიგადის ნომერი", "თარიღი", "ID",
    "ობიექტის დასახელება", "მისამართი", "ქალაქი კოეფიციენტი (გადაადგილება)", "შესრულებული სამუშაო",
    "ბრიგადის წევრთა რაოდენობა", "დაწყება", "დასრულება", "რაოდენობა ჯამი",
    "რაოდენობა კაცზე", "კომენტარი", "შენიშვნა", "თანამდებობა"
]

COL = {name: i + 1 for i, name in enumerate([
    "city", "name", "position_id", "brigade", "date", "id",
    "object_name", "address", "coefficient", "work_type",
    "member_count", "start", "end", "qty_total",
    "qty_per_person", "comment", "note", "position"
])}

LEADER_POSITIONS = {"ბრიგადირი", "სარემონტო ბრიგადის უფროსი"}


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def get_db_connection():
    conn = psycopg2.connect(
        os.environ.get("POSTGRES_URL"),
        cursor_factory=RealDictCursor
    )
    return conn


def init_db():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS records (
                db_id VARCHAR(100) PRIMARY KEY,
                id VARCHAR(100) NOT NULL,
                data JSONB NOT NULL,
                created_at DATE NOT NULL
            );
        """)
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print("DB Init Error:", e)


# აპლიკაციის სტარტზე იქმნება ბაზა (თუ არ არსებობს)
try:
    init_db()
except Exception as e:
    print("Startup DB init skipped/failed:", e)


def generate_excel_from_records(records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "შესრულებული სამუშაოები"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    row_num = 2

    for record in records:
        id_val = record.get("id", "")
        brigade = record.get("brigade", "")
        city = record.get("city", "")
        date = record.get("date", "")
        object_name = record.get("object_name", "")
        address = record.get("address", "")
        coefficient = record.get("coefficient", "")
        overall_comment = record.get("overall_comment", "")
        member_data = record.get("members", [])
        works = record.get("works", [])

        active_members, absent_members = [], []
        for m in member_data:
            note = (m.get("note") or "").strip()
            entry = {
                "name": m["name"],
                "position": m["position"],
                "personal_id": m.get("personal_id", ""),
                "extra": bool(m.get("extra")),
                "home_brigade": m.get("home_brigade"),
            }
            if m.get("absent", False) or note:
                entry["note"] = note or "არ გამოცხადება"
                absent_members.append(entry)
            else:
                active_members.append(entry)

        # იერარქიული სორტირება: ბრიგადირი პირველ ადგილზე
        active_members.sort(key=lambda x: 0 if x.get("position") in LEADER_POSITIONS else 1)
        absent_members.sort(key=lambda x: 0 if x.get("position") in LEADER_POSITIONS else 1)

        worker_count = sum(1 for m in active_members if m["position"] not in LEADER_POSITIONS)

        def position_id_label(member):
            label = f"{member['position']} {member['personal_id']}".strip()
            if member.get("extra"):
                label += f" (ბრიგადა {member.get('home_brigade', '')}-დან)"
            return label

        def member_count_value(member):
            return 1 if member["position"] in LEADER_POSITIONS else worker_count

        for work in works:
            work_type = work.get("work_type", "")
            start_val = work.get("start", "")
            end_val = work.get("end", "")
            try:
                total_qty = float(work.get("quantity", 0) or 0)
            except (TypeError, ValueError):
                total_qty = 0
            work_comment = work.get("comment", "")

            for member in active_members:
                divisor = member_count_value(member)
                per_person = round(total_qty / divisor, 2) if divisor else 0
                row = [""] * len(HEADERS)
                row[COL["city"] - 1] = city
                row[COL["name"] - 1] = member["name"]
                row[COL["position_id"] - 1] = position_id_label(member)
                row[COL["brigade"] - 1] = brigade
                row[COL["date"] - 1] = date
                row[COL["id"] - 1] = id_val
                row[COL["object_name"] - 1] = object_name
                row[COL["address"] - 1] = address
                row[COL["coefficient"] - 1] = coefficient
                row[COL["work_type"] - 1] = work_type
                row[COL["member_count"] - 1] = member_count_value(member)
                row[COL["start"] - 1] = start_val
                row[COL["end"] - 1] = end_val
                row[COL["qty_total"] - 1] = total_qty
                row[COL["qty_per_person"] - 1] = per_person
                row[COL["comment"] - 1] = work_comment if work_comment else overall_comment
                row[COL["note"] - 1] = ""
                row[COL["position"] - 1] = member["position"]
                ws.append(row)
                row_num += 1

        for member in absent_members:
            row = [""] * len(HEADERS)
            row[COL["city"] - 1] = city
            row[COL["name"] - 1] = member["name"]
            row[COL["position_id"] - 1] = position_id_label(member)
            row[COL["brigade"] - 1] = brigade
            row[COL["date"] - 1] = date
            row[COL["id"] - 1] = id_val
            row[COL["object_name"] - 1] = object_name
            row[COL["address"] - 1] = address
            row[COL["coefficient"] - 1] = coefficient
            row[COL["work_type"] - 1] = "არ გამოცხადება"
            row[COL["member_count"] - 1] = 0
            row[COL["start"] - 1] = ""  # სრულიად ცარიელი ველი "დაწყება"-სთვის
            row[COL["end"] - 1] = ""    # სრულიად ცარიელი ველი "დასრულება"-სთვის
            row[COL["note"] - 1] = member["note"]
            row[COL["position"] - 1] = member["position"]
            ws.append(row)
            row_num += 1

    widths = [14, 22, 26, 12, 12, 10, 20, 20, 12, 30, 12, 10, 10, 12, 12, 20, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------- ძირითადი MARŞRUTები ----------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/get_brigades")
def get_brigades():
    config = load_config()
    brigades = sorted(config["brigades"].keys(), key=lambda x: int(x))
    return jsonify(brigades)


@app.route("/get_brigade_cities")
def get_brigade_cities():
    brigade = request.args.get("brigade")
    if not brigade:
        return jsonify([])
    config = load_config()
    members = config["brigades"].get(brigade, {}).get("members", [])
    cities = sorted({m["city"] for m in members if "city" in m})
    return jsonify(cities)


@app.route("/get_brigade_members")
def get_brigade_members():
    brigade = request.args.get("brigade")
    city = request.args.get("city", "")
    if not brigade:
        return jsonify([])
    config = load_config()
    members = config["brigades"].get(brigade, {}).get("members", [])
    if city:
        members = [m for m in members if m.get("city") == city]
    return jsonify(members)


@app.route("/get_all_members")
def get_all_members():
    config = load_config()
    all_members = []
    for brigade_num, data in config["brigades"].items():
        for m in data.get("members", []):
            all_members.append({
                "name": m["name"],
                "position": m["position"],
                "personal_id": m.get("personal_id", ""),
                "city": m.get("city", ""),
                "brigade": brigade_num,
            })
    all_members.sort(key=lambda m: (int(m["brigade"]), m["name"]))
    return jsonify(all_members)


@app.route("/get_work_types")
def get_work_types():
    config = load_config()
    return jsonify(config.get("work_types", []))


# ---------- POSTGRES API-ები ----------

@app.route("/api/records", methods=["GET"])
def api_get_records():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT data, created_at FROM records ORDER BY created_at DESC;")
        rows = cur.fetchall()
        cur.close()
        conn.close()

        records = [r["data"] for r in rows]
        
        has_old = False
        now = datetime.now().date()
        for r in rows:
            if (now - r["created_at"]).days >= 365:
                has_old = True
                break

        return jsonify({"records": records, "has_old_records": has_old})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/records/save", methods=["POST"])
def api_save_record():
    data = request.get_json(silent=True)
    if not data or "id" not in data:
        return jsonify({"error": "არასწორი მონაცემები"}), 400

    rec_id = str(data["id"])
    db_id = str(data.get("db_id") or f"{rec_id}_{int(datetime.now().timestamp()*1000)}")
    data["db_id"] = db_id
    rec_date = data.get("date", datetime.now().strftime("%Y-%m-%d"))

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO records (db_id, id, data, created_at)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (db_id) DO UPDATE SET data = EXCLUDED.data, created_at = EXCLUDED.created_at;
        """, (db_id, rec_id, json.dumps(data, ensure_ascii=False), rec_date))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/records/delete/<path:record_id>", methods=["DELETE"])
def api_delete_record(record_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM records WHERE db_id = %s OR id = %s;", (str(record_id), str(record_id)))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/generate", methods=["POST"])
def generate():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "მონაცემები არასწორია"}), 400

    records = data.get("records", [])
    if not records:
        return jsonify({"error": "ჩანაწერები ვერ მოიძებნა"}), 400

    for rec in records:
        for field in ("id", "brigade", "city", "date", "address"):
            if not rec.get(field):
                return jsonify({"error": f"აკლია სავალდებულო ველი: {field}"}), 400
        if not rec.get("works"):
            return jsonify({"error": f"ჩანაწერს ID {rec.get('id')} არ აქვს სამუშაოები"}), 400

    try:
        output = generate_excel_from_records(records)
    except Exception as exc:
        return jsonify({"error": f"Excel-ის გენერირება ვერ მოხერხდა: {exc}"}), 500

    filename = f"ანგარიში_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
